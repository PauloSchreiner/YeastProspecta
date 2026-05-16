import pandas as pd
import numpy as np
from Bio import SearchIO, SeqIO

import re
import os 

from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill



#
# EXTRACTION
#

def parse_blast_xmls(xml_path_list: list) -> pd.DataFrame:
    """
    Reads BLAST XML files and returns a DataFrame 
    containing all hits across all samples.
    """
    rows = []

    for xml_file in xml_path_list:
        sample = os.path.basename(xml_file).rsplit("_", 1)[0] 
        qresult = SearchIO.read(xml_file, "blast-xml")

        for hit_index, hit in enumerate(qresult, start=1):
            hsp = hit[0] 

            qseq = str(hsp.query.seq)
            hseq = str(hsp.hit.seq)

            # Extract indel events
            indels, subs, MNV_positions = parse_evo_events(qseq, hseq)

            MNV_positions = str(MNV_positions).strip('[]') # Transform list into string and remove []

            rows.append({
                "Sample": sample.zfill(3),
                "Hit #": hit_index,
                "Accession": hit.accession,
                "Identity": round((hsp.ident_num / hsp.aln_span) * 100, 3),
                "Subst": subs,
                "Indels": indels,
                "MNV positions": MNV_positions,
                "Coverage": round((hsp.query_span / qresult.seq_len) * 100, 3),
                "E-value": hsp.evalue,
                "Title": hit.description
            })

    return pd.DataFrame(rows)

def extract_ambiguity_data(fasta_path_list: list) -> pd.DataFrame:
    """
    Reads the consensus FASTA files and counts the number of ambiguous 
    bases (IUPAC codes and Ns) to detect mixed cultures or intragenomic variation.
    """
    rows = []
    for fasta in fasta_path_list:
        sample = os.path.basename(fasta).split(".")[0].replace("_consensus", "").zfill(3)
        
        # Read the FASTA
        record = list(SeqIO.parse(fasta, "fasta"))[0]
        seq = str(record.seq).upper()
        
        # Count anything that is not A, T, C, G or - (gap)
        ambiguous_count = len(re.findall(r'[^ATCG-]', seq))
        
        rows.append({"Sample_ID": sample, "Ambiguous_bases": ambiguous_count})
        
    df = pd.DataFrame(rows)
    df.set_index("Sample_ID", inplace=True)
    return df


def extract_trim_data(trim_path: str) -> pd.DataFrame:
    """
    Reads the trimming CSV, pivots Forward (F) and Reverse (R) reads into separate columns, 
    and returns a DataFrame indexed by sample.
    """
    # 1. Read the entire file at once (sep='\t' indicates it's Tab-separated)
    df = pd.read_csv(trim_path, sep="\t")
    
    # 2. Split the 'sample' column (e.g., "1_F") into two new columns: ID and Direction.
    # expand=True creates actual DataFrame columns.
    df[['Sample_ID', 'Direction']] = df['sample'].str.rsplit('_', n=1, expand=True)
    
    # 3. Standardize sample names using zfill(3) (e.g., "1" becomes "001") so we can join tables later.
    df['Sample_ID'] = df['Sample_ID'].astype(str).str.zfill(3)
    
    # 4. Pivot the table so 'F' and 'R' directions become their own columns
    df_pivot = df.pivot(index='Sample_ID', columns='Direction', values='post_trim_length')
    
    # 5. Rename columns to our desired standard
    df_pivot.rename(columns={'F': 'F_len', 'R': 'R_len'}, inplace=True)
    
    # Clean up the column axis name for aesthetics and return
    df_pivot.columns.name = None 
    return df_pivot


def extract_consensus_data(consensus_path: str) -> pd.DataFrame:
    """
    Reads the consensus CSV, standardizes sample names, and safely handles missing or invalid scores.
    """
    df = pd.read_csv(consensus_path, sep="\t")
    
    # 1. Rename columns to match our standard naming convention
    df.rename(columns={
        'Sample': 'Sample_ID', 
        'Length': 'Consensus_len', 
        'Score': 'Consensus_score'
    }, inplace=True)
    
    # 2. Standardize sample names (e.g., "1" -> "001") for flawless merging
    df['Sample_ID'] = df['Sample_ID'].astype(str).str.zfill(3)
    
    # 3. Safely cast the Score column to numeric. 'coerce' turns errors/blanks into NaNs, 
    # which we then fill with 0 and cast to integers.
    df['Consensus_score'] = pd.to_numeric(df['Consensus_score'], errors='coerce').fillna(0).astype(int)
    
    # 4. Set the Sample_ID as the DataFrame index (the row labels)
    df.set_index('Sample_ID', inplace=True)
    
    return df



#
# PROCESSING
#


def extract_species_from_title(ncbi_title:str) -> str:
    """
    Extracts the binomial species name from a raw NCBI title.
    
    Args: ncbi_title: The raw title from BLAST.
    Returns: The clean 'Genus species' string.
    """

    pattern = r"^(\[?[A-Z][a-z]+\]?\s+[a-z]+\.?)"
    match = re.search(pattern, ncbi_title)

    if match:
        return match.group(1)
    else: 
        return ncbi_title


def is_type(ncbi_title:str) -> bool:
    """
    Given an NCBI title, checks if it is the type strain.
    """
    return ("type" in ncbi_title.lower())



def parse_evo_events(qseq: str, hseq: str) -> (int, int, list):
  """
  Based on a previously aligned pair of query sequence and hit sequence,
  returns three values: the amount of indel events (non-contiguous), 
  the amount of substitution events (non-contiguous)
  and the positions of MNVs (Multiple-Nucleotide Variants, AKA contiguous indels or substitutions)
  """
  
  indel_events = subst_events = 0
  in_subst_block = in_indel_block = False
  MNV_pos = []

  position = 0
  for qchar, hchar in zip(qseq, hseq):
    
    # Indel
    if qchar == "-" or hchar == "-": 
      if in_indel_block: 
        MNV_pos.append(position)
      else: # If this is the first time seeing this block
        indel_events += 1
      in_indel_block = True 

    # Substitution
    elif qchar != hchar: 
      if in_subst_block:
        MNV_pos.append(position)
      else: # If this is the first time seeing this block
        subst_events += 1
      in_subst_block = True
    
    # Match (no subst or indel)
    else:
      in_subst_block = in_indel_block = False
    
    # Updates position 
    position += 1

  return indel_events, subst_events, MNV_pos



def build_qc_df(df_trim: pd.DataFrame, df_cons: pd.DataFrame, df_blast: pd.DataFrame, df_ambig: pd.DataFrame, id_thresh: float) -> pd.DataFrame:
    """
    Joins Trimming and Consensus data, and calculates the number of hits above the identity threshold from the BLAST results.
    Returns the consolidated Quality Control DataFrame ready for export.
    """
    # Join Trim and Consensus data, as well as ambiguity data. 
    # An 'outer' join ensures we don't lose samples that might be missing.
    df_qc = df_trim.join(df_cons, how='outer')
    df_qc = df_qc.join(df_ambig, how='outer')
    
    # Calculate how many hits are above id threshold
    hits_above_thresh = df_blast[df_blast['Identity'] > id_thresh].groupby('Sample').size()
    
    # Name this new Series so it becomes a properly named column after joining
    hits_above_thresh.name = 'Hits_above_' + str(id_thresh) + '_id'
    
    # Add the hits_above_thresh count to the main QC DataFrame
    df_qc = df_qc.join(hits_above_thresh, how='outer')
    
    # NaN handling and Data Typing (Avoiding the "Float curse")
    # The Pandas 'Int64' type (capital 'I') allows integers to coexist with NaNs without converting the whole column to decimals.
    colunas_numericas = ['F_len', 'R_len', 'Consensus_len', 'Consensus_score', 'Ambiguous_bases', hits_above_thresh.name]
    
    for col in colunas_numericas:
        if col in df_qc.columns:
            # Fill NaNs with 0 and force the safe integer type
            df_qc[col] = df_qc[col].fillna(0).astype('Int64')
            
    # Remove extra index naming and sort alphabetically by Sample
    df_qc.index.name = "Sample"
    df_qc.sort_index(inplace=True)
            
    return df_qc



def build_summary_df(df_qc: pd.DataFrame, df_blast: pd.DataFrame, id_thresh: float, 
                     good_cons_score: int, min_len: int, min_cov: float, max_ambig: int) -> pd.DataFrame:
    """
    Generates the summary report by cross-referencing QC data with the Top Hits from BLAST.
    """
    
    # 1. Isolamos TODAS as cepas tipo
    is_type_mask = df_blast["Title"].apply(is_type)
    df_all_type = df_blast[is_type_mask].copy()

    # Extrai o nome limpo da espécie
    df_all_type['Species'] = df_all_type['Title'].apply(extract_species_from_title)
    df_all_type['Species'] = df_all_type['Species'].str.replace(r'[\[\]]', '', regex=True)
    
    # Ordena para garantir que o melhor hit absoluto de tipo fique no topo
    df_all_type.sort_values(by=['Sample', 'Identity'], ascending=[True, False], inplace=True)
    
    # O df_top agora captura o tipo mais próximo da história, mesmo que seja baixo!
    df_top = df_all_type.drop_duplicates(subset=['Sample'], keep='first').copy()
    df_top = df_top[['Sample', 'Species', 'Identity', 'Accession', 'Coverage', 'Subst', 'Indels', 'MNV positions']].rename(
        columns={
            'Species': 'Top hit (type)',
            'Identity': 'Identity (type)',
            'Accession': 'Accession (type)',
            'Coverage': 'Coverage (type)',
            'Subst': 'Subst events',
            'Indels': 'Indel events',
            'MNV positions' : 'MNV positions'
        }
    )

    # 2. Para a coluna de ambiguidades (Other possible species), só queremos os que competem ACIMA do threshold
    df_type_above = df_all_type[df_all_type["Identity"] > id_thresh].copy()
    
    df_unique_species = df_type_above.drop_duplicates(subset=['Sample', 'Species'], keep='first').copy()
    df_unique_species['Formatted'] = df_unique_species['Species'] + " (" + df_unique_species['Identity'].astype(str) + "%)"
    
    # Agrupa pulando o primeiro hit (iloc[1:]), pois ele já é o nosso Top Hit
    df_other = df_unique_species.groupby('Sample').apply(
        lambda x: "; ".join(x['Formatted'].iloc[1:])
    ).reset_index(name='Other possible species')

    df_highest_id = df_blast.groupby("Sample")["Identity"].max().rename("Highest identity overall")
    
    # --- MERGING EVERYTHING ---
    
    # Name has to match the one determined in build_qc_df
    hits_above_thresh_name = 'Hits_above_' + str(id_thresh) + '_id'

    # Get the baseline QC data needed for the final rules
    df_final = df_qc.reset_index()[['Sample', 'Consensus_score', 'Consensus_len','Ambiguous_bases', hits_above_thresh_name]].copy()
    
    # Merge (Left Join) the Top Hits and the Other Species back into our baseline
    df_final = df_final.merge(df_top, on='Sample', how='left')
    df_final = df_final.merge(df_other, on='Sample', how='left')
    df_final = df_final.merge(df_highest_id, on='Sample', how='left')
    
    # Fill blanks for samples that had NO hits passing above threshold or Type Strain filters
    df_final['Top hit (type)'] = df_final['Top hit (type)'].fillna("None")
    df_final['Identity (type)'] = df_final['Identity (type)'].fillna(0)
    df_final['Accession (type)'] = df_final['Accession (type)'].fillna("N/A")
    df_final['Coverage (type)'] = df_final['Coverage (type)'].fillna(0)
    df_final['Subst events'] = df_final['Subst events'].fillna(0).astype(int)
    df_final['Indel events'] = df_final['Indel events'].fillna(0).astype(int)
    df_final['MNV positions'] = df_final['MNV positions'].fillna("")
    df_final['Highest identity overall'] = df_final['Highest identity overall'].fillna(0) 
    df_final['Other possible species'] = df_final['Other possible species'].fillna("")
    
    
    # --- BUSINESS RULES (STATUS EVALUATION) ---
    # np.select evaluates multiple conditions simultaneously, replacing if/elif/else loops!
    conditions = [
        df_final['Consensus_score'] < good_cons_score,      # 1. Bad quality
        df_final['Consensus_len'] < min_len,                # 2. Short sequence
        df_final['Ambiguous_bases'] > max_ambig,            # 3. Ambiguous bases
        df_final['Coverage (type)'] < min_cov,              # 4. Low coverage 
        df_final['Highest identity overall'] < id_thresh,   # 5. No hit above threshold
        df_final['Identity (type)'] < id_thresh,            # 6. No hit with type
        (df_final['Subst events'] + df_final['Indel events']) >= 4, # 7. 4 or more mutations
        df_final['Other possible species'] != ""            # 8. Many species 
    ]
    
    labels = [
        "Bad quality",
        "Short sequence",
        "Ambiguous bases",
        "Low coverage",
        "No hit above threshold",
        "No hit with type",
        "4 or more mutations", 
        "Many species"
    ]
        
    # Applies the conditions in order. If none match, defaults to "OK".
    df_final['Status'] = np.select(conditions, labels, default="OK")
    
    # Reorder the final columns and return
    colunas_finais = ["Sample", 
                      "Status", 
                      "Top hit (type)", 
                      "Identity (type)", 
                      "Coverage (type)",
                      "Accession (type)",
                      "Consensus_len",
                      "Consensus_score",
                      "Subst events", 
                      "Indel events", 
                      "MNV positions", 
                      "Highest identity overall", 
                      "Other possible species"]
    
    return df_final[colunas_finais]



#
# PRESENTING
#


def auto_adjust_columns(writer: pd.ExcelWriter, sheet_name: str):
    """
    Adjusts the width of all columns in the specified Excel sheet 
    to fit the content.
    """
    worksheet = writer.sheets[sheet_name] 
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter # Get 'A', 'B', etc.
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Capped at 60 characters so really long descriptions don't break the UI
        adjusted_width = min(max_length + 2, 60)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def apply_workbook_styling(writer: pd.ExcelWriter, df_summary: pd.DataFrame, df_blast: pd.DataFrame, 
                           df_qc: pd.DataFrame, good_cons_score: int, max_ambig: int, 
                           min_len: int, min_cov: float, id_thresh: float):
    """
    Applies all conditional formatting and styling to the entire Excel workbook.
    """

    colors = {
    "green":        PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid'), 
    "lime":         PatternFill(start_color='E9F5BC', end_color='E9F5BC', fill_type='solid'), 
    "yellow":       PatternFill(start_color='FFF3BF', end_color='FFF3BF', fill_type='solid'), 
    
    "blue":         PatternFill(start_color='D1ECF1', end_color='D1ECF1', fill_type='solid'), 
    "red":          PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid'), 
    "purple":       PatternFill(start_color='EADDFF', end_color='EADDFF', fill_type='solid'), 
    "orange":       PatternFill(start_color='FFD8B1', end_color='FFD8B1', fill_type='solid'),
    
    "light_gray":   PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'), 
    "white":        PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'), 

    "gray":         PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'), 
    "brown":        PatternFill(start_color='E6D5C3', end_color='E6D5C3', fill_type='solid'),
    
    "event_blue":   PatternFill(start_color='CFE2FF', end_color='CFE2FF', fill_type='solid'), 
    "event_purple": PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid'), 
    "mnv_red":      PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')  
}

    # --- SHEET 1: SUMMARY (Status Colors) ---
    ws_sum = writer.sheets["Summary"]
    status_range = f"B2:B{len(df_summary) + 1}"
    
    ws_sum.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"OK"'], fill=colors["green"]))
    ws_sum.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"Bad quality"'], fill=colors["red"]))
    ws_sum.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"Many species"'], fill=colors["yellow"]))
    ws_sum.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"Noisy consensus"'], fill=colors["orange"]))
    ws_sum.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"No hit with type"'], fill=colors["blue"]))
    ws_sum.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"No hit above threshold"'], fill=colors["purple"]))
    ws_sum.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"Short sequence"'], fill=colors["gray"]))
    ws_sum.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"Low coverage"'], fill=colors["brown"]))
    ws_sum.conditional_formatting.add(status_range, CellIsRule(operator='equal', formula=['"4 or more mutations"'], fill=colors["event_purple"]))

    # --- Mapeamento e Regras de Métricas Técnicas e Taxonômicas ---
    col_map = {cell.value: cell.column_letter for cell in ws_sum[1]}
    
    len_col = col_map.get("Consensus_len")
    score_sum_col = col_map.get("Consensus_score")
    cov_col = col_map.get("Coverage (type)")
    id_type_col = col_map.get("Identity (type)")
    highest_id_col = col_map.get("Highest identity overall")
    
    last_row = len(df_summary) + 1

    # 1. Comprimento do Consenso abaixo do mínimo (Vermelho)
    if len_col:
        len_range = f"{len_col}2:{len_col}{last_row}"
        ws_sum.conditional_formatting.add(len_range, CellIsRule(operator='lessThan', formula=[str(min_len)], fill=colors["red"]))

    # 2. Score de Alinhamento abaixo do ideal (Vermelho)
    if score_sum_col:
        score_sum_range = f"{score_sum_col}2:{score_sum_col}{last_row}"
        ws_sum.conditional_formatting.add(score_sum_range, CellIsRule(operator='lessThan', formula=[str(0.75 * good_cons_score)], fill=colors["red"]))

    # 3. Cobertura de Alinhamento abaixo do mínimo (Vermelho)
    if cov_col:
        cov_range = f"{cov_col}2:{cov_col}{last_row}"
        ws_sum.conditional_formatting.add(cov_range, CellIsRule(operator='lessThan', formula=[str(min_cov)], fill=colors["red"]))

    # 4. Identidade com a Cepa Tipo abaixo do threshold (Vermelho)
    if id_type_col:
        id_type_range = f"{id_type_col}2:{id_type_col}{last_row}"
        
        # Abaixo do mínimo (Vermelho)
        ws_sum.conditional_formatting.add(id_type_range, CellIsRule(operator='lessThan', formula=[str(id_thresh)], fill=colors["red"]))
        # Entre o mínimo e 99.999% (Azul)
        ws_sum.conditional_formatting.add(id_type_range, CellIsRule(operator='between', formula=[str(id_thresh), '99.999'], fill=colors["blue"]))
        # Exatamente 100% (Verde)
        ws_sum.conditional_formatting.add(id_type_range, CellIsRule(operator='equal', formula=['100'], fill=colors["green"]))

    # 5. Alerta de Discrepância / Espécie Nova Críptica (Amarelo)
    if highest_id_col:
        highest_id_range = f"{highest_id_col}2:{highest_id_col}{last_row}"
        
        # Abaixo do mínimo (Vermelho)
        ws_sum.conditional_formatting.add(highest_id_range, CellIsRule(operator='lessThan', formula=[str(id_thresh)], fill=colors["red"]))
        # Entre o mínimo e 99.999% (Azul)
        ws_sum.conditional_formatting.add(highest_id_range, CellIsRule(operator='between', formula=[str(id_thresh), '99.999'], fill=colors["blue"]))
        # Exatamente 100% (Verde)
        ws_sum.conditional_formatting.add(highest_id_range, CellIsRule(operator='equal', formula=['100'], fill=colors["green"]))

        # Alerta de Discrepância / Cepa Nova (Amarelo)
        if id_type_col:
            ws_sum.conditional_formatting.add(highest_id_range, FormulaRule(
                formula=[f"=AND(${highest_id_col}2>=99.0, ${id_type_col}2<{id_thresh})"], 
                fill=colors["yellow"]
            ))

    # 6. Color coding evolutionary events (indels and subst) and MNVs
    col_map = {cell.value: cell.column_letter for cell in ws_sum[1]}
    sub_col = col_map.get("Subst events")
    ind_col = col_map.get("Indel events")
    mnv_col = col_map.get("MNV positions")

    last_row = len(df_summary) + 1
    if sub_col:
        sub_range = f"{sub_col}2:{sub_col}{last_row}"
        # >= 6 (Roxo)
        ws_sum.conditional_formatting.add(sub_range, CellIsRule(operator='greaterThanOrEqual', formula=['6'], fill=colors["event_purple"]))
        # Entre 4 e 5 (Azul)
        ws_sum.conditional_formatting.add(sub_range, CellIsRule(operator='between', formula=['4', '5'], fill=colors["event_blue"]))

    if ind_col:
        ind_range = f"{ind_col}2:{ind_col}{last_row}"
        # Qualquer Indel > 0 (Roxo, indicando mutação estrutural relevante)
        ws_sum.conditional_formatting.add(ind_range, CellIsRule(operator='greaterThan', formula=['0'], fill=colors["event_purple"]))
    
    if mnv_col:
        mnv_range = f"{mnv_col}2:{mnv_col}{last_row}"
        ws_sum.conditional_formatting.add(mnv_range, 
            FormulaRule(formula=[f'LEN(TRIM(${mnv_col}2))>0'], fill=colors["mnv_red"]))

    # --- SHEET 2: BLAST DETAILS (Alternating Sample Groups) ---
    ws_blast = writer.sheets["BLAST_Details"]
    current_fill = colors["white"]
    last_sample = None
    
    for i, sample_id in enumerate(df_blast["Sample"], start=2):
        if last_sample is not None and sample_id != last_sample:
            current_fill = colors["light_gray"] if current_fill == colors["white"] else colors["white"]
        
        for col in range(1, ws_blast.max_column + 1):
            ws_blast.cell(row=i, column=col).fill = current_fill
        last_sample = sample_id

    # --- SHEET 3: QUALITY CONTROL (Consensus Score Rules) ---
    ws_qc = writer.sheets["Quality_Control"]
    
    # Identify the 'Consensus_score' column dynamically
    score_col_letter = None
    for cell in ws_qc[1]:
        if cell.value == "Consensus_score":
            score_col_letter = cell.column_letter
            break
    
    ambig_col_letter = None
    for cell in ws_qc[1]:
        if cell.value == "Ambiguous_bases":
            ambig_col_letter = cell.column_letter
            break
            
    if ambig_col_letter:
        qc_ambig_range = f"{ambig_col_letter}2:{ambig_col_letter}{len(df_qc) + 1}"
        ws_qc.conditional_formatting.add(qc_ambig_range, CellIsRule(operator='greaterThan', formula=[str(max_ambig)], fill=colors["orange"]))
        
    
    if score_col_letter:
        qc_range = f"{score_col_letter}2:{score_col_letter}{len(df_qc) + 1}"
        ws_qc.conditional_formatting.add(qc_range, CellIsRule(operator='greaterThan', formula=[str(good_cons_score * 1.2)], fill=colors["green"]))
        ws_qc.conditional_formatting.add(qc_range, CellIsRule(operator='greaterThan', formula=[str(good_cons_score)], fill=colors["lime"]))
        ws_qc.conditional_formatting.add(qc_range, CellIsRule(operator='greaterThan', formula=[str(good_cons_score * 0.8)], fill=colors["yellow"]))
        ws_qc.conditional_formatting.add(qc_range, CellIsRule(operator='lessThanOrEqual', formula=[str(good_cons_score * 0.8)], fill=colors["red"]))


def export_to_excel(df_summary: pd.DataFrame, df_blast: pd.DataFrame, df_qc: pd.DataFrame, 
                    output_file: str, max_excel_hits: int, good_cons_score: int, 
                    max_ambig: int, min_len: int, min_cov: float, id_thresh: float):
    """
    Orchestrates the export of DataFrames to Excel sheets, 
    applies visual filters (Top X hits), and triggers formatting.
    """
    # 1. Presentation Filter: Keep only the Top X hits in the BLAST sheet
    df_blast_export = df_blast[df_blast['Hit #'] <= max_excel_hits].copy()

    # 2. Final aesthetic sorting
    df_summary.sort_values(by="Sample", inplace=True)
    df_blast_export.sort_values(by=["Sample", "Hit #"], inplace=True)
    
    # Prepare df_qc by turning the 'Sample' index into a real column so Excel displays it properly
    df_qc_export = df_qc.reset_index() 

    # 3. Write data to the Excel file
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        
        # Write DataFrames to their respective sheets
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_blast_export.to_excel(writer, sheet_name="BLAST_Details", index=False)
        df_qc_export.to_excel(writer, sheet_name="Quality_Control", index=False)

        # --- FREEZE TOP ROWS ---
        # Iterate over all created sheets and freeze the top row (A2 means freeze everything above row 2)
        for sheet_name in writer.sheets:
            writer.sheets[sheet_name].freeze_panes = "A2"
        # -------------------------------------------

        # Auto-adjust column widths
        auto_adjust_columns(writer, "Summary")
        auto_adjust_columns(writer, "BLAST_Details")
        auto_adjust_columns(writer, "Quality_Control")

        # Apply conditional formatting based on the already filtered df_blast
        apply_workbook_styling(writer, df_summary, df_blast_export, df_qc_export, 
                               good_cons_score, max_ambig, min_len, min_cov, id_thresh)
#
# MAIN LOGIC
#

def main():
    # Get parameters from snakefile
    id_thresh = float(snakemake.params.id_thresh)
    max_excel_hits = int(snakemake.params.max_excel_hits)
    good_consensus_score = int(snakemake.params.good_consensus_score)
    min_len = int(snakemake.params.min_consensus_len)
    min_cov = float(snakemake.params.min_coverage)
    max_ambig = int(snakemake.params.max_ambiguous_bases)

    # --- 1. EXTRACTION ---
    df_trim = extract_trim_data(snakemake.input.trim_summary)
    df_cons = extract_consensus_data(snakemake.input.consensus_summary)
    df_ambig = extract_ambiguity_data(snakemake.input.fastas) 
    df_blast_completo = parse_blast_xmls(snakemake.input.xmls)
    
    # --- 2. PROCESSING ---
    df_qc = build_qc_df(df_trim, df_cons, df_blast_completo, df_ambig, id_thresh)
    df_summary = build_summary_df(df_qc, 
                                  df_blast_completo, 
                                  id_thresh, 
                                  good_consensus_score, 
                                  min_len, 
                                  min_cov,
                                  max_ambig)
    
    # --- 3. PRESENTATION / EXPORTING ---
    export_to_excel(df_summary, df_blast_completo, df_qc, snakemake.output.report, 
                    max_excel_hits, good_consensus_score, max_ambig, min_len, min_cov, id_thresh)


if __name__ == "__main__":
    main()